# -*- coding: utf-8 -*-
'''
@FileName	:   fsel.py
@Created     :   2021/02/24 13:58
@Updated    :   2022/08/28 13:58
@Author		:   Teddy, goonhope@gmail.com, Zhuhai
@Function	:   selenium 二次封装Template
'''

import os, time, xlwt
from selenium import webdriver
from PIL import Image
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
import selenium.webdriver.support.expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import Select
from urllib.parse import urljoin
from Project.function import str_chk,fdir,timer
from time import sleep
from random import uniform


def err(func):
    """装饰器：错误时返回函数名称"""
    def inner(self,*args,**kwargs):
        goal = func(self,*args,**kwargs)
        print(f"check {func.__name__}!") if not goal else None
        return goal
    return inner


def iter_click(ifs=False,css="a",ink="",slp=1.12):
    """循环点击"""
    def inner(func):
        def inr(self,*args,**kwargs):
            tags = [x for x in self.driver.find_elements("css selector",css) if x.text and str_chk(x.text,ink=ink)]
            for en,tag in enumerate(tags):
                tag = tags[en]
                tag.click()
                self.tsleep(1,slp)
                if ifs: break
                hold = func(self,*args,**kwargs)
                self.driver.back()
                self.tsleep(1,slp)
                tags = [x for x in self.driver.find_elements("css selector",css)  if x.text and str_chk(x.text,ink=ink)]
            return hold
        return inr
    return inner


class CDriver(object):
    """selenium二次封装"""
    def __init__(self, driver,wrong=False):
        self.driver = driver
        self.wrong = wrong
        if wrong:  # 网站证书问题时，默认正常
            self.click_by_css("button#details-button")
            self.click_by_css("a#proceed-link")
            # self.driver.switch_to.alert.accept() if EC.alert_is_present() else None
            # accept-确定、dismiss-取消

    def init_page(self,url):
        self.driver.get(url)
        self.driver.implicitly_wait(2)
        # self.driver.maximize_window()
        # self.driver.set_page_load_timeout(10)
        # self.driver.set_script_timeout(10)

    def tsleep(self,min=1.,max=0.):
        """固定随机sleep时间"""
        sleep(uniform(min, max)) if max > min else sleep(min)

    def js(self,jstr="alert(\"请输入验证码\\n停留时间8s\");",slp=1,accept=False):
        '''执行js命令'''
        result = self.driver.execute_script(jstr)
        self.tsleep(1)
        if accept: self.driver.switch_to.alert.accept()
        self.tsleep(slp)
        return result

    def show(self,css,timeout=5, gone=False):
        '''检测等待css是否显示或消失'''
        located = EC.visibility_of_element_located(("css selector", css))
        try:
            ui.WebDriverWait(self.driver, timeout).until(located) if not gone else \
            ui.WebDriverWait(self.driver, timeout).until_not(located)
            return True
        except TimeoutException:
            return False

    def is_visible(self,css):
        '''是否显示'''
        return self.driver.find_element("css selector",css).is_displayed()

    def options_select(self,css,order=0,des=False):
        '''下拉菜单选择_CSS为选项直接上级'''
        elem = Select(self.driver.find_element("css selector",css))
        elem.select_by_index(order) if isinstance(order,int) else elem.select_by_value(order)
        elem.deselect_all() if elem.is_multiple and des else None # 取消所有选择

    def input_key(self,ids,css="input.el-input__inner",n=0,clear=False):
        """输入数据"""
        if isinstance(ids,str):
            input_el = self.driver.find_elements("css selector",css)[n]
            input_el.clear()
            input_el.send_keys(ids)
        else:
            _ids = len(ids)
            input_el = self.driver.find_elements("css selector",css)[:_ids]
            if _ids == len(input_el):
                for x,i in zip(input_el,ids):
                    ActionChains(self.driver).double_click(x).perform()
                    if clear: x.clear()
                    x.send_keys(i)
            else:
                print("check input_key function")
        self.tsleep(0.5)

    def click_by_css(self, css="li.el-menu-item.pull-right a",n=0,slp=3,cs=True):
        """点击css 默认css 点击，否则js点击"""
        if cs:
            self.driver.find_elements("css selector",css)[n].click()
            self.tsleep(2,slp)
        else:
            s = f"document.querySelectorAll('{css}')[{n}].click()"
            self.js(s, slp, False)
        return

    def iter_click(self,func,ifs=False,css="a",ink="",slp=1.12,*args,**kwargs):
        """循环点击"""
        tags = [x for x in self.driver.find_elements("css selector",css) if x.text and str_chk(x.text,ink=ink)]
        for en,tag in enumerate(tags):
            tag = tags[en]
            tag.click()
            self.tsleep(1,slp)
            if ifs: break
            hold = func(*args,**kwargs)
            self.driver.back()
            self.tsleep(1,slp)
            tags = [x for x in self.driver.find_elements("css selector",css)  if x.text and str_chk(x.text,ink=ink)]
        return hold

    def grap_sub(self,tag,subcss, by="css selector"):
        """获取子节点"""
        childrens_tag = tag.find_elements(by,subcss)
        return childrens_tag

    def grap_tag(self, css, by="css selector",n=True):
        '''获取tag'''
        tag = self.driver.find_elements(by,css) if n else self.driver.find_element(by,css)
        return tag if tag else print("check css!") or False

    def choose_tag(self,css,by="css selector",**kwargs):
        "选择标签"
        tags = [x for x in self.driver.find_elements(by,css) if x.text and str_chk(x.text,**kwargs)]
        return tags if tags else False

    def tag_content(self,css,tag="text",by="css selector",**kwargs):
        "获取tag内容 -单一-数组"
        content = [x.get_attribute(tag) if tag != "text" else x.text for x in self.driver.find_elements(by,css) if x.text and str_chk(x.text,**kwargs)]
        return content if content else False

    def get_content(self,css,tag="text",by="css selector"):
        "获取tag内容 -单一-数组"
        content = [x.get_attribute(tag) if tag != "text" else x.text for x in self.driver.find_elements(by,css)]
        return content if content else False

    @timer()
    def tag_contents(self,css,tag="text",by="css selector",**kwargs):
        "获取tag内容-多属性-字典数组 "
        fn = lambda tag, j: {i :j.get_attribute(i) if i != "text" else j.text for i in tag.strip().split() if j.get_attribute(i)}
        content = [fn(tag,x) for x in self.driver.find_elements(by,css) if x.text and str_chk(x.text,**kwargs)]
        return content if content else False

    def get_css_attr(self,css="[id^=srollmap]",plus=True):
        ''''获取css所有属性及值, return [dict,]'''
        jstr = f"""return Object.values(document.querySelectorAll("{css}")).map(x=>Object.values(x.attributes).map(i=>i.name + ":" + i.nodeValue));""" \
                    if plus else f""" return Object.values(document.querySelector("{css}").attributes).map(i=>i.name + ":" + i.nodeValue);"""
        kv = self.js(jstr,accept=False)
        kv = kv if plus else [kv]
        kv = [{x[:x.find(":")].strip(): x[x.find(":") + 1:].strip() for x in i} for i in kv]
        return kv

    def shot(self, pic,css="#cms_r",fdir=""):
        '''截图'''
        fdir = fdir or os.path.dirname(__file__)   # 默认文件所在目录
        img_file = os.path.join(os.path.join(fdir, f"{pic}_full.png"))
        self.driver.get_screenshot_as_file(img_file)  # 获取网页截图
        elem = self.driver.find_element("css selector",css)
        left,top = [int(elem.location[x]) for x in "xy"]
        right = int(left + elem.size['width'])
        bottom = int(top + elem.size['height'])
        im = Image.open(img_file)
        im = im.crop((left, top, right, bottom))  # 获取css元素大小截图
        fo = os.path.join(fdir, f"{pic}.png")
        im.save(os.path.join(fo))
        return os.path.exists(fo)

    def _quit(self,d=True):
        """默认退出，否则关闭当前页面"""
        self.driver.quit() if d else self.driver.close()

    def switch(self,page=1):
        """切换页面"""
        print(self.driver.window_handles)
        page = page if page <= len(self.driver.window_handles) else 0
        page = self.driver.window_handles[page]
        self.driver.switch_to.window(page)


def move_file(dirname,dn,fn,subdir="",show=True):
    ''''移动到指定目录'''
    dir_default = r"D:\Downloads"  # chrome 默认下载目录
    subdir = subdir or r"D:\Temp\{}".format(dirname) # 移动到目录
    os.mkdir(subdir) if not os.path.exists(subdir) else None
    fdn = os.path.join(dir_default, dn)  # 下载文件全名
    sleep(3) if not os.path.exists(fdn) else None  # 等待文件下载
    ffn = os.path.join(subdir, fn)
    os.renames(fdn,ffn)
    os.system("start %s" % subdir) if show else None
    return True if os.path.exists(ffn) else False


@timer()
def crm(url="",show=True,proxy=None,user=False,img=True,ddir=""):
    """chrome, options- don't reload image"""
    option, prefs = webdriver.ChromeOptions(), {"plugins.always_open_pdf_externally": True}
    if user:
        userprofile = os.popen("echo %tmp%").read().strip()
        userprofile = userprofile.strip("Temp") + r"Google\Chrome\User Data\Default"
        option.add_argument(f'user-data-dir={userprofile}')
    if not img: prefs.update({"profile.managed_default_content_settings.images": 2}) # 不加载图片
    if ddir and os.path.exists(ddir): prefs.update({"download.default_directory": ddir}) # 默认下载目录
    option.add_experimental_option('prefs',prefs)
    option.add_experimental_option('excludeSwitches', ['enable-automation'])
    if not show: option.add_argument('--headless') # 显示
    if proxy: option.add_argument(f"--proxy-server={proxy.split('://')[-1] if '://' in proxy else proxy}")
    option.add_argument('--incongnito') # 隐身默认
    option.add_argument("--ignore-certificate-errors")
    driver = webdriver.Chrome(options=option)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",{
    "source":"""Object.defineProperty(nabigator,'webdriver',{get: () => undefined })"""}) # 置空
    driver.maximize_window()
    driver.get(url) if url else None
    # driver.implicitly_wait(5)
    # print(option._argumentsx.quit, option._experimental_options)
    return driver


def get_cookies(url,**kwargs):
    """获取cookies"""
    driver = crm(url,**kwargs)
    x = driver.execute_script("""return localStorage""")
    data = driver.get_cookies()
    driver.close()
    return data,x


def fox(url="",show=True,proxy=None,img=True,ddir=""):
    """Fox"""
    option = webdriver.FirefoxOptions()
    if proxy: option.add_argument(f"--proxy-server={proxy.split('://')[-1] if '://' in proxy else proxy}")
    option.add_argument('headless') if not show else None
    option.set_preference("permissions.default.image", 2) if img else None
    if ddir and os.path.exists(ddir):
        option.set_preference('browser.download.dir', ddir)
        option.set_preference('browser.download.folderList', 2)
    driver = webdriver.Firefox(options=option)
    driver.maximize_window()
    driver.get(url) if url else None
    return driver


def excelin(info,file_out,rc=True):
    '''直接写入excel,info二维数组 [[]]'''
    file = xlwt.Workbook()
    table = file.add_sheet('list', cell_overwrite_ok=True)
    style = xlwt.XFStyle()
    for row, rows in enumerate(info): # 行 y-->Row
        for col, cell in enumerate(rows): # 列 x--> Column
            table.write(row, col, cell, style) if rc else table.write(col, row, cell, style)
    file.save(file_out)


def rgb(row):
    d, m = divmod(row, 256)
    sig = hex(m).replace("0x", "").zfill(2)
    rgb = sig *3
    return rgb


def excel_raw(xlsx,info=None,sheet=0,rg=(0,0,0,0),color=""):
    '''读取写入xlsx 源文件直接公式'''
    from openpyxl import Workbook,load_workbook
    from openpyxl.styles import Font,PatternFill
    if info: # 写入
        wb = Workbook()
        table = wb.create_sheet('list',0)
        if not color:
            for x in info: table.append(x)
        else:
            for row, rows in enumerate(info,1):  # 行 y-->Row
                for col, data in enumerate(rows,1):  # 列 x--> Column
                    table.cell(row, col).fill = PatternFill("solid", fgColor=rgb(row))
                    table.cell(row, col, value=data)
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        wb.save(xlsx)
        return
    else: # 读取
        wb = load_workbook(xlsx,read_only=True)
        sheet = wb.worksheets[sheet]
        if isinstance(rg,tuple):
            min_col, min_row, max_col, max_row = rg
            chose = sheet.iter_rows(*rg)
        elif isinstance(rg, str) and ":" in rg:
            chose = sheet[rg] # rg = "c6:d11" 选区
        else: chose = sheet.rows
        data =[[cell.value for cell in rs] for rs in chose]
        return data


def excel(func):
    '''写入excel装饰器'''
    def wrapper(*args, **kwargs):
        start = time.time()
        info,company = func(*args, **kwargs)
        file_out = os.path.join(fdir(), f'{company}_备案信息_{time.strftime("%Y%m%d_%H%M%S")}.xls')
        excelin(info,file_out)
        print(f"@{func.__name__}:\t[Time:{time.time() - start : 0.3f}s]")
        return
    return wrapper


def stamp_from(date_str,fmt="%Y%m%d"):
    # 转换为时间戳:
    timeArray = time.strptime(date_str,fmt)
    timeStamp = int(time.mktime(timeArray))
    return timeStamp


def time_from(stamp,fmt="%Y%m%d %H:%M:%S"):
    """时间戳转字符串"""
    return time.strftime(fmt, time.localtime(stamp))


def swap_time(st,fmt="%Y%m%d %H:%M:%S"):
    """时间戳与时间字符串互转"""
    if isinstance(st,str):return stamp_from(st,fmt)
    elif isinstance(st,(int,float)): return time_from(st,fmt)
    else:print("check st")


@timer()
def chrome_():
    d = CDriver(crm(r"http://ssthjj.zhuhai.gov.cn/zxfw/xmgsgg/slgg/index.html",img=False))
    print(d.tag_contents("a","text href",ink="报告"))
    print(d.tag_content("a", ink="报告"))
    d._quit()


if __name__ == '__main__':
    # xlsx = os.path.join(fdir(),"x.xlsx")
    # xlsx = r"C:\Users\surface\Downloads\b.xlsx"
    # print(excel_raw(xlsx,rg=(3,36,11,46)))
    chrome_()
